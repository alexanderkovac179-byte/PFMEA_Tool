import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.legends import (
    VYZNAM_LEGENDA,
    VYSKYT_LEGENDA,
    ODHALENIE_LEGENDA,
    KLASIFIKACIA_LEGENDA,
    ISTOTA_LEGENDA,
)
from app.config import FMEA_METADATA_DEFAULTS


# ── Farby pre klasifikáciu (CC / OS / SC / HI) ────────────────────────────────
# Používa sa v legende aj ako referencia pre farebné kódovanie
KLASIFIKACIA_FILLS = {
    "CC": PatternFill(fill_type="solid", fgColor="F8CBAD"),  # oranžovo-červená
    "OS": PatternFill(fill_type="solid", fgColor="FFE699"),  # žltá
    "SC": PatternFill(fill_type="solid", fgColor="C6E0B4"),  # zelená
    "HI": PatternFill(fill_type="solid", fgColor="BDD7EE"),  # modrá
}


def _estimate_row_height(
    ws,
    row_idx: int,
    widths: dict,
    min_height: int = 48,
    max_height: int = 240,
) -> int:
    """
    Odhad potrebnej výšky riadku pre wrap_text bunky.

    Excel nemá spoľahlivý auto-height pre multi-line content, preto počítame:
      approx_chars_per_line = column_width * 1.3   (Calibri 10pt, empiricky)
      needed_lines          = ceil(text_len / chars_per_line)
      row_height            = needed_lines * 15pt + padding

    Vracia hodnotu v points (1pt ≈ 1.33px) v intervale [min_height, max_height].
    """
    max_lines = 1
    for col_letter, width in widths.items():
        val = ws[f"{col_letter}{row_idx}"].value
        if val is None or val == "":
            continue

        text = str(val)
        chars_per_line = max(int(width * 1.3), 6)

        lines_in_cell = 0
        # Rešpektuj explicitné zalomenia (\n) a k nim pridaj wrap lines.
        for paragraph in text.split("\n"):
            if not paragraph:
                lines_in_cell += 1
                continue
            lines_in_cell += (len(paragraph) + chars_per_line - 1) // chars_per_line
        max_lines = max(max_lines, lines_in_cell)

    # 15pt na riadok textu + ~6pt padding
    height = max_lines * 15 + 6
    return max(min_height, min(height, max_height))


def write_legend_sheet(ws, data, confidence_colors=False, klasifikacia_colors=False,
                       severity_colors=False, occurrence_colors=False, detection_colors=False):
    thin        = Side(style="thin",   color="000000")
    medium_side = Side(style="medium", color="000000")
    thin_border   = Border(left=thin,        right=thin,        top=thin,        bottom=thin)
    medium_border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    bold        = Font(bold=True)
    bold10      = Font(bold=True, size=10)
    bold12      = Font(bold=True, size=12)

    # Farby pre Istota AI legendu
    conf_fills = {
        "Vysoká istota":        PatternFill(fill_type="solid", fgColor="D9EAD3"),  # zelená
        "Stredná istota":       PatternFill(fill_type="solid", fgColor="FFF2CC"),  # žltá
        "Nízka istota":         PatternFill(fill_type="solid", fgColor="F4CCCC"),  # červená
        "Dôležité upozornenie": PatternFill(fill_type="solid", fgColor="F3F3F3"),  # šedá
    }

    # Farebný gradient pre Severity / Occurrence / Detection (hodnoty 1-10)
    # Zelená (nízke riziko) → žltá (stredné) → oranžová → červená (vysoké)
    def _risk_gradient(index: int) -> str:
        """Vráti hex farbu pre hodnotu indexu 1-10."""
        gradient = {
            10: "C00000",  # sýto červená
            9:  "E06666",  # červená
            8:  "F4B084",  # oranžová
            7:  "F8CBAD",  # svetlo-oranžová
            6:  "FFD966",  # oranžovo-žltá
            5:  "FFE699",  # svetlá žltá
            4:  "E2EFDA",  # svetlozelená
            3:  "C6E0B4",  # zelená
            2:  "A9D08E",  # tmavšie zelená
            1:  "70AD47",  # sýta zelená
        }
        return gradient.get(index, "FFFFFF")

    # Pre legendy Vyznam/Vyskyt/Odhalenie – detekuj hodnotu indexu v poslednom
    # alebo predposlednom stĺpci (môže byť napr. 10 alebo "9 – 10")
    def _parse_index(value):
        """Z hodnoty legendy extrahuje číselný index 1-10 alebo None."""
        if isinstance(value, int):
            return value if 1 <= value <= 10 else None
        if isinstance(value, str):
            # "9 – 10" → 9 (berieme najvyššiu hodnotu rozsahu)
            parts = [p.strip() for p in value.replace("–", "-").split("-")]
            try:
                numbers = [int(p) for p in parts if p.isdigit()]
                if numbers:
                    return max(numbers)
            except (ValueError, TypeError):
                pass
        return None

    # Šedá farba pre "Dôležité upozornenie" aj v legende klasifikácie
    note_fill = PatternFill(fill_type="solid", fgColor="F3F3F3")

    # Ktorý stĺpec obsahuje "hodnotu indexu" – posledný v riadku
    index_col_idx = len(data[0]) if data else 0

    for r_idx, row in enumerate(data, start=1):
        row_label = str(row[0]).strip() if row else ""

        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border

            if r_idx == 1:
                # Hlavička – svetlomodrý podklad, bold, centrovanie
                cell.fill      = header_fill
                cell.font      = Font(bold=True, size=11)
                cell.alignment = center
                cell.border    = medium_border

            else:
                # Dátové riadky – všetko na stred
                cell.alignment = center

                # Confidence legenda – farebné riadky
                if confidence_colors and row_label in conf_fills:
                    cell.fill = conf_fills[row_label]
                    if c_idx in (1, 2):
                        cell.font = Font(bold=True, size=11)
                    else:
                        cell.font = Font(size=11)
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True)

                # Legendy Vyznam / Vyskyt / Odhalenie – gradient podľa indexu
                elif severity_colors or occurrence_colors or detection_colors:
                    # Riadok je poznámka? (originálna alebo o norme)
                    is_note = row_label.lower().startswith("poznámka")
                    # Posledný stĺpec má hodnotu indexu – zoberieme ju
                    index_value = row[-1] if row else None
                    idx = _parse_index(index_value)
                    if idx is not None and not is_note:
                        cell.fill = PatternFill(
                            fill_type="solid",
                            fgColor=_risk_gradient(idx),
                        )
                        cell.font = (Font(bold=True, size=11) if c_idx == 1
                                     else Font(size=11))
                    elif is_note:
                        cell.fill = note_fill
                        # Prvý stĺpec (názov "Poznámka o norme") – tučné
                        # Ostatné stĺpce – kurzíva (vysvetľujúci text)
                        if c_idx == 1:
                            cell.font = Font(bold=True, size=11)
                        else:
                            cell.font = Font(italic=True, size=11)
                    else:
                        cell.font = Font(size=11)

                # Legenda klasifikácie – farebná len prvá bunka (skratka)
                elif klasifikacia_colors:
                    if c_idx == 1:
                        if row_label in KLASIFIKACIA_FILLS:
                            cell.fill = KLASIFIKACIA_FILLS[row_label]
                            cell.font = bold12
                        elif row_label == "Dôležité upozornenie":
                            cell.fill = note_fill
                            cell.font = Font(bold=True, size=11)
                    elif row_label == "Dôležité upozornenie":
                        # celý riadok upozornenia na šedom pozadí, kurzíva
                        cell.fill = note_fill
                        cell.font = Font(italic=True, size=11)
                    else:
                        cell.font = Font(size=11)

                else:
                    # Štandardné legendy – jednotná veľkosť 11
                    cell.font = Font(size=11)
                    if isinstance(value, (int, float)) or str(value).strip() in {"X", "x"}:
                        cell.alignment = center
                    else:
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True)

    # Šírky stĺpcov – širšie aby sa zmestili dlhé texty poznámok
    for col in ws.columns:
        max_length  = 0
        col_letter  = col[0].column_letter
        for cell in col:
            if cell.value:
                # Všetky riadky textu – berieme najdlhší
                for line in str(cell.value).split("\n"):
                    max_length = max(max_length, len(line))
        # Širší rozsah: 12-65 znakov; pri kratších stĺpcoch min šírka 18
        col_width = max(18, min(max_length + 4, 65))
        ws.column_dimensions[col_letter].width = col_width

    # Výška riadkov – dynamicky podľa najdlhšieho textu v riadku
    ws.row_dimensions[1].height = 30
    for row_idx, row in enumerate(data, start=1):
        if row_idx == 1:
            continue
        # Najdlhší text v riadku
        max_len = 0
        for value in row:
            if value is None:
                continue
            text = str(value).strip()
            if text and text != "—":
                max_len = max(max_len, len(text))
        # Heuristika: ~50 znakov na riadok textu, 18 px na riadok
        # Min 60 px, max 240 px
        approx_lines = max(2, (max_len // 50) + 1)
        height = max(60, min(approx_lines * 22, 240))
        ws.row_dimensions[row_idx].height = height

    # Print layout pre legendy – landscape, fit na šírku, okraje
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:1"
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top  = ws.page_margins.bottom = 0.5
    ws.print_options.horizontalCentered = True
    ws.oddFooter.center.text = "Strana &P z &N"
    ws.oddFooter.center.size = 9
    ws.oddHeader.left.text = "&A"   # názov hárku (Legenda_…)
    ws.oddHeader.left.size = 10
    ws.oddHeader.right.text = "&D"
    ws.oddHeader.right.size = 9


def _append_odhalenie_note(ws):
    """
    Pod tabuľku Legenda_Odhalenie pridá vysvetlenie stĺpcov
    "Typ kontroly A / B / C". Hodnoty prichádzajú z AIAG/VDA PFMEA.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin        = Side(style="thin",   color="000000")
    medium_side = Side(style="medium", color="000000")
    thin_border   = Border(left=thin,        right=thin,        top=thin,        bottom=thin)
    medium_border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side)

    note_fill   = PatternFill(fill_type="solid", fgColor="F3F3F3")
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    bold10      = Font(bold=True, size=10)
    left_wrap   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 2 riadky medzery pod poslednou existujúcou hodnotou
    start_row = ws.max_row + 2
    last_col  = ws.max_column

    # ── Nadpis poznámky (merged cez všetky stĺpce) ───────────────────────────
    ws.cell(row=start_row, column=1, value="Vysvetlenie typov kontroly")
    ws.merge_cells(
        start_row=start_row, start_column=1,
        end_row=start_row, end_column=last_col,
    )
    header_cell = ws.cell(row=start_row, column=1)
    header_cell.font      = Font(bold=True, size=11)
    header_cell.fill      = header_fill
    header_cell.alignment = center
    header_cell.border    = medium_border
    # Ostatné bunky v merged rozsahu tiež dostanú border (openpyxl to inak nemerguje)
    for c in range(2, last_col + 1):
        ws.cell(row=start_row, column=c).border = medium_border

    ws.row_dimensions[start_row].height = 24

    # ── 3 riadky popisu A / B / C ────────────────────────────────────────────
    rows = [
        ("A", "Kontrola zabraňujúca chybe",
         "Konštrukčné alebo technické riešenie, ktoré fyzicky zabraňuje vzniku "
         "chyby – napr. poka-yoke, mistake proofing, jednoznačné polohovanie, "
         "geometrické vylúčenie, blokovanie procesu pri nezhode."),
        ("B", "Meranie hodnoty / parametra",
         "Kontrola kvantitatívneho parametra – kalibrované meradlá, SPC, "
         "automatické meranie momentu, rozmerov, tlaku, elektrických veličín, "
         "kamerové systémy s merateľným výstupom."),
        ("C", "Manuálna / vizuálna kontrola",
         "Kontrola operátorom alebo kontrolórom – vizuálne overenie, "
         "porovnanie s etalónom, čítanie záznamu, OK/NOK rozhodnutie "
         "na základe zmyslov alebo jednoduchých pomôcok."),
    ]

    for offset, (label, title, desc) in enumerate(rows):
        r = start_row + 1 + offset

        # Bunka A/B/C – centrovaná, tučná
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font      = Font(bold=True, size=14)
        c1.fill      = note_fill
        c1.alignment = center
        c1.border    = thin_border

        # Nadpis typu kontroly (2. stĺpec)
        c2 = ws.cell(row=r, column=2, value=title)
        c2.font      = bold10
        c2.fill      = note_fill
        c2.alignment = left_wrap
        c2.border    = thin_border

        # Popis (ostatné stĺpce merged)
        ws.cell(row=r, column=3, value=desc)
        ws.merge_cells(
            start_row=r, start_column=3,
            end_row=r, end_column=last_col,
        )
        c3 = ws.cell(row=r, column=3)
        c3.font      = Font(size=10)
        c3.fill      = note_fill
        c3.alignment = left_wrap
        c3.border    = thin_border
        for c in range(4, last_col + 1):
            ws.cell(row=r, column=c).fill   = note_fill
            ws.cell(row=r, column=c).border = thin_border

        ws.row_dimensions[r].height = 48


def export_to_excel(items: list[dict], output_path: str, metadata: dict | None = None,
                   include_legends: bool = True):
    metadata = {**FMEA_METADATA_DEFAULTS, **(metadata or {})}

    desired_order = [
        "funkcia_procesu_pozadavky",
        "mozna_chyba",
        "mozny_nasledok_chyby",
        "vyznam",
        "klasifikacia",
        "mozna_pricina_mechanizmus_chyby",
        "vyskyt",
        "pouzivane_metody_prevencie",
        "pouzivane_metody_odhalenia",
        "odhalenie",
        "rpn",
        "doporucene_opatrenia",
        "zodp_pracovnik_datum_ukoncenia",
        "novy_vyznam",
        "novy_vyskyt",
        "nove_odhalenie",
        "novy_rpn",
        "confidence"
    ]

    rename_map = {
        "funkcia_procesu_pozadavky": "Funkcia procesu / požiadavky",
        "mozna_chyba": "Možná chyba",
        "mozny_nasledok_chyby": "Možný následok chyby",
        "vyznam": "Význam",
        "klasifikacia": "Klasifikácia",
        "mozna_pricina_mechanizmus_chyby": "Možná príčina / mechanizmus chyby",
        "vyskyt": "Výskyt",
        "pouzivane_metody_prevencie": "Používané metódy k prevencii voči výskytu",
        "pouzivane_metody_odhalenia": "Používané metódy k odhaleniu",
        "odhalenie": "Odhalenie",
        "rpn": "RPN",
        "doporucene_opatrenia": "Doporučené opatrenia",
        "zodp_pracovnik_datum_ukoncenia": "Zodp. pracovník / dátum ukončenia",
        "novy_vyznam": "Nový význam",
        "novy_vyskyt": "Nový výskyt",
        "nove_odhalenie": "Nové odhalenie",
        "novy_rpn": "Nový RPN",
        "confidence": "Istota AI"
    }

    df = pd.DataFrame(items)

    # Confidence: float → percentá pre zobrazenie v Exceli
    if "confidence" in df.columns:
        df["confidence"] = df["confidence"].apply(
            lambda x: round(float(x) * 100) if x not in ("", None) else ""
        )

    for col in desired_order:
        if col not in df.columns:
            df[col] = ""

    df = df[desired_order]
    df = df.sort_values(by="rpn", ascending=False)
    df = df.rename(columns=rename_map)

    # startrow=7 → pandas zapíše stĺpcové hlavičky do riadka 8 a dátové riadky
    # od riadka 9. V pôvodnom kóde bolo startrow=6, pri čom pandas dával dáta
    # od riadka 8 a následný "copy row 7 → row 8" prepisoval prvý dátový riadok
    # (každý export ticho stratil jednu položku s najvyšším RPN).
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Procesna_FMEA", startrow=7)

    wb = load_workbook(output_path)
    ws = wb["Procesna_FMEA"]

    thin = Side(style="thin", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    medium_border = thin_border

    title_fill = PatternFill(fill_type="solid", fgColor="9DC3E6")
    info_label_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    group_fill = PatternFill(fill_type="solid", fgColor="BDD7EE")
    header_fill = PatternFill(fill_type="solid", fgColor="EAF2F8")

    red_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    green_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    normal_font = Font(size=10)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ── Hlavička: titulok "Procesná FMEA" roztiahnutý až po stĺpec R
    #    (predtým A1:Q1 → vizuálna "diera" na pravej strane nad stĺpcom
    #    "Istota AI (%)")
    ws.merge_cells("A1:R1")
    ws["A1"] = metadata["nazov_hlavicky"]
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center
    ws["A1"].border = medium_border

    today = datetime.now().strftime("%d.%m.%Y")

    info_rows = {
        "A2": "Názov procesu:",
        "F2": metadata["nazov_procesu"],
        "J2": "Číslo FMEA:",
        "N2": metadata["cislo_fmea"],

        "A3": "Typ FMEA:",
        "F3": metadata["typ_fmea"],
        "J3": "Revízia:",
        "N3": metadata["revizia"],

        "A4": "Vypracoval:",
        "F4": metadata["vypracoval"],
        "J4": "Dátum:",
        "N4": today,

        "A5": "Poznámka:",
        "F5": metadata["poznamka"],
    }

    # Poznámka v riadku 5 teraz končí pri R (predtým Q) — rovnaký dôvod ako pri
    # titulku v riadku 1.
    merge_ranges = [
        "A2:E2", "F2:I2", "J2:M2", "N2:R2",
        "A3:E3", "F3:I3", "J3:M3", "N3:R3",
        "A4:E4", "F4:I4", "J4:M4", "N4:R4",
        "A5:E5", "F5:R5",
    ]
    for r in merge_ranges:
        ws.merge_cells(r)

    for ref, value in info_rows.items():
        ws[ref] = value

    # Info bloky pokrývame po stĺpec R (18) aby col R mal rámček aj v riadkoch 2-5.
    for row in ws.iter_rows(min_row=2, max_row=5, min_col=1, max_col=18):
        for cell in row:
            cell.border = thin_border
            cell.alignment = left
            cell.font = normal_font

    for cell in ["A2", "A3", "A4", "A5", "J2", "J3", "J4"]:
        ws[cell].font = bold
        ws[cell].fill = info_label_fill

    group_row = 7
    header_row = 8
    data_start_row = 9

    # Pandas už zapísal stĺpcové hlavičky do riadka 8 (header_row) a dáta od
    # riadka 9 (data_start_row). Riadok 7 (group_row) je prázdny a pripravený
    # na zlúčené skupinové hlavičky definované nižšie v `groups`.
    # Druhé úrovňové sub-hlavičky (Prevencia / Odhalenie / Nový význam ...)
    # prepíšu príslušné bunky riadka 8 cez `second_headers`.

    groups = [
        ("A7:A8", "Funkcia"),
        ("B7:B8", "Možná chyba"),
        ("C7:C8", "Možný následok chyby"),
        ("D7:D8", "Význam"),
        ("E7:E8", "Klasifikácia"),
        ("F7:F8", "Možná príčina / mechanizmus chyby"),
        ("G7:G8", "Výskyt"),
        ("H7:I7", "Súčasné riadenie procesu"),
        ("J7:J8", "Odhalenie"),
        ("K7:K8", "RPN"),
        ("L7:L8", "Doporučené opatrenia"),
        ("M7:M8", "Zodp. pracovník / dátum ukončenia"),
        ("N7:Q7", "Výsledky opatrení"),
        ("R7:R8", "Istota AI (%)"),
    ]

    for rng, value in groups:
        ws.merge_cells(rng)
        first = ws[rng.split(":")[0]]
        first.value = value
        first.font = bold
        first.fill = group_fill
        first.alignment = center
        first.border = medium_border

    second_headers = {
        "H8": "Prevencia",
        "I8": "Odhalenie",
        "N8": "Nový význam",
        "O8": "Nový výskyt",
        "P8": "Nové odhalenie",
        "Q8": "Nový RPN",
    }

    for ref, value in second_headers.items():
        ws[ref] = value

    # Rozšírené na max_col=18 aby R7/R8 boli korektne nastavené
    for row in ws.iter_rows(min_row=7, max_row=8, min_col=1, max_col=18):
        for cell in row:
            cell.font = bold
            if cell.fill.fill_type is None:
                cell.fill = header_fill
            cell.alignment = center
            cell.border = medium_border

    # Rovnako pre dátové riadky – stĺpec R (Istota AI) teraz dostane borders
    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, min_col=1, max_col=18):
        for cell in row:
            cell.alignment = left
            cell.border = thin_border
            cell.font = normal_font

    # Klasifikácia – stĺpec E: centrovanie a farebné kódovanie
    # (CC = oranžová, OS = žltá, SC = zelená, HI = modrá)
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    for row in range(data_start_row, ws.max_row + 1):
        cell = ws[f"E{row}"]
        val  = str(cell.value or "").strip().upper()
        cell.alignment = center
        if val in KLASIFIKACIA_FILLS:
            cell.fill = KLASIFIKACIA_FILLS[val]
            cell.font = Font(bold=True, size=10)
        else:
            cell.fill = white_fill

    numeric_cols = ["D", "E", "G", "J", "K", "N", "O", "P", "Q", "R"]
    for col in numeric_cols:
        for row in range(data_start_row, ws.max_row + 1):
            ws[f"{col}{row}"].alignment = center

    widths = {
        "A": 34,
        "B": 30,
        "C": 36,
        "D": 10,
        "E": 18,
        "F": 36,
        "G": 10,
        "H": 34,
        "I": 34,
        "J": 10,
        "K": 10,
        "L": 38,
        "M": 24,
        "N": 12,
        "O": 12,
        "P": 14,
        "Q": 12,
        "R": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[7].height = 24
    ws.row_dimensions[8].height = 34

    # Dynamická výška dátových riadkov – namiesto fixných 64 bodov počítame
    # potrebnú výšku podľa najdlhšieho textu v riadku vzhľadom na šírku stĺpca.
    # Text v stĺpcoch C / L (Doporučené opatrenia) má často 200+ znakov;
    # fixných 64 px ich orezalo. Strop je EXCEL_ROW_HEIGHT_MAX aby sa výška
    # neodpútala pri výnimočne dlhom obsahu.
    from app.config import EXCEL_ROW_HEIGHT_MIN, EXCEL_ROW_HEIGHT_MAX

    for row in range(data_start_row, ws.max_row + 1):
        ws.row_dimensions[row].height = _estimate_row_height(
            ws, row, widths,
            min_height=EXCEL_ROW_HEIGHT_MIN,
            max_height=EXCEL_ROW_HEIGHT_MAX,
        )

    for row in range(data_start_row, ws.max_row + 1):
        cell = ws[f"K{row}"]
        try:
            val = int(cell.value)
            if val >= 200:
                cell.fill = red_fill
            elif val >= 100:
                cell.fill = yellow_fill
            else:
                cell.fill = green_fill
            cell.font = bold
            cell.alignment = center
        except Exception:
            pass

    for row in range(data_start_row, ws.max_row + 1):
        cell = ws[f"Q{row}"]
        try:
            val = int(cell.value)
            if val >= 200:
                cell.fill = red_fill
            elif val >= 100:
                cell.fill = yellow_fill
            else:
                cell.fill = green_fill
            cell.font = bold
            cell.alignment = center
        except Exception:
            pass

    # Confidence – podmienené formátovanie stĺpca R
    conf_high_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")   # zelená ≥ 75 %
    conf_mid_fill  = PatternFill(fill_type="solid", fgColor="FFF2CC")   # žltá  50–74 %
    conf_low_fill  = PatternFill(fill_type="solid", fgColor="F4CCCC")   # červená < 50 %

    for row in range(data_start_row, ws.max_row + 1):
        cell = ws[f"R{row}"]
        cell.alignment = center
        cell.font = Font(bold=True, size=10)
        try:
            val = int(cell.value)
            if val >= 75:
                cell.fill = conf_high_fill
            elif val >= 50:
                cell.fill = conf_mid_fill
            else:
                cell.fill = conf_low_fill
            cell.value = f"{val} %"
        except (TypeError, ValueError):
            pass

    for col in range(1, 19):
        ws.cell(7, col).border = medium_border
        ws.cell(ws.max_row, col).border = medium_border

    for row in range(7, ws.max_row + 1):
        ws.cell(row, 1).border = medium_border
        ws.cell(row, 18).border = medium_border

    ws.auto_filter.ref = f"A8:R{ws.max_row}"
    ws.freeze_panes = "A9"

    # ── Print layout – aby PDF a tlač vyzerali rozumne ─────────────────────────
    # Orientácia landscape, fit na jednu stranu šírky, opakovanie hlavičky
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0   # šírka sa musí zmestiť, výšku môže prelievať
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    # Opakuj riadky 1–8 (metadata + hlavička) na každej strane pri tlači
    ws.print_title_rows = "1:8"
    # Print area – od A1 po R{last} (nepridávaj prázdne stĺpce)
    ws.print_area = f"A1:R{ws.max_row}"
    # Okraje a centrovanie na stranu
    ws.page_margins.left   = 0.4
    ws.page_margins.right  = 0.4
    ws.page_margins.top    = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.print_options.horizontalCentered = True
    # Päta s číslovaním strán
    ws.oddFooter.center.text = "Strana &P z &N"
    ws.oddFooter.center.size = 9
    ws.oddFooter.center.color = "666666"
    # Záhlavie tlače s názvom
    ws.oddHeader.left.text = "Procesná FMEA"
    ws.oddHeader.left.size = 10
    ws.oddHeader.left.color = "1F4E79"
    ws.oddHeader.right.text = "&D"   # automatický dátum tlače
    ws.oddHeader.right.size = 9
    ws.oddHeader.right.color = "666666"

    # ── Legendy ────────────────────────────────────────────────────────────────
    # Nanovo vytvárame, aby sa pri opakovanom spustení neduplikovali.
    # Pri include_legends=False sa všetky existujúce legendy odstránia
    # a žiadne nové sa nevytvoria – výstup je len FMEA tabuľka.
    legend_sheets = [
        "Legenda_Vyznam",
        "Legenda_Vyskyt",
        "Legenda_Odhalenie",
        "Legenda_Klasifikacia",
        "Legenda_Istota_AI",
    ]
    for sheet_name in legend_sheets:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    if include_legends:
        ws_vyznam = wb.create_sheet("Legenda_Vyznam")
        write_legend_sheet(ws_vyznam, VYZNAM_LEGENDA, severity_colors=True)

        ws_vyskyt = wb.create_sheet("Legenda_Vyskyt")
        write_legend_sheet(ws_vyskyt, VYSKYT_LEGENDA, occurrence_colors=True)

        ws_odhalenie = wb.create_sheet("Legenda_Odhalenie")
        write_legend_sheet(ws_odhalenie, ODHALENIE_LEGENDA, detection_colors=True)
        _append_odhalenie_note(ws_odhalenie)

        ws_klasifikacia = wb.create_sheet("Legenda_Klasifikacia")
        write_legend_sheet(ws_klasifikacia, KLASIFIKACIA_LEGENDA, klasifikacia_colors=True)

        ws_istota = wb.create_sheet("Legenda_Istota_AI")
        write_legend_sheet(ws_istota, ISTOTA_LEGENDA, confidence_colors=True)

    wb.save(output_path) 