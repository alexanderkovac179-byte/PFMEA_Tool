"""
cost_estimator.py – odhad ceny a kvality vstupných dát pre PFMEA Tool.

Modul poskytuje dve funkcie pre PreflightDialog:

1. estimate_cost()  – odhad ceny API volaní pre Claude Opus 4.7
                      na základe veľkosti vstupných dokumentov.

2. analyze_quality() – analyzuje typy dokumentov a detekuje, či obsahujú
                       sekcie typické pre kvalitnú FMEA (kontrolný plán,
                       záznamy o nezhodách, pracovné postupy).

Tokenizácia: pre slovenčinu sa používa empirická konštanta 2.5 znaku/token
(vychádza z faktu, že slovenský text spotrebuje ~1.7× viac tokenov ako
anglický, kde platí 4 znaky/token).
"""

from pathlib import Path
from dataclasses import dataclass


# ── Cenová tabuľka (USD za 1 milión tokenov) ───────────────────────────────────
# Aktuálne k apríl 2026, stanica https://www.anthropic.com/pricing
MODEL_PRICING = {
    "claude-opus-4-7":      {"input": 5.00,  "output": 25.00},
    "claude-opus-4-6":      {"input": 5.00,  "output": 25.00},
    "claude-sonnet-4-6":    {"input": 3.00,  "output": 15.00},
    "claude-haiku-4-5":     {"input": 0.80,  "output":  4.00},
    "claude-haiku-4-5-20251001": {"input": 0.80, "output": 4.00},
}

# Default ak model nie je v tabuľke (Opus ako bezpečný horný odhad)
DEFAULT_PRICING = {"input": 5.00, "output": 25.00}

# Empirická konštanta pre slovenský text:
# 1 token ≈ 2.5 znakov (pre angličtinu je to ~4 znaky)
CHARS_PER_TOKEN_SLOVAK = 2.5

# Odhad output tokenov na FMEA položku (7 polí v slovenčine)
TOKENS_PER_FMEA_ITEM = 500

# Odhadovaný počet FMEA položiek na krok
ITEMS_PER_STEP_AVG = 2.5

# Odhadovaný počet krokov na 1000 znakov vstupu
STEPS_PER_1K_CHARS = 0.4

# Overhead system promptu + JSON schémy v každom volaní
SYSTEM_PROMPT_TOKENS = 600


@dataclass
class CostEstimate:
    """Výsledok odhadu nákladov na jeden beh PFMEA Tool."""
    total_input_chars:  int
    total_input_tokens: int
    estimated_steps:    int
    estimated_items:    int
    output_tokens:      int
    cost_input_usd:     float
    cost_output_usd:    float
    cost_total_usd:     float
    cost_total_eur:     float
    model:              str
    files_count:        int

    def format_summary(self) -> str:
        """Človeku čitateľný súhrn pre zobrazenie v UI."""
        return (
            f"Odhadovaná cena: ${self.cost_total_usd:.2f} "
            f"(≈ {self.cost_total_eur:.2f} €)"
        )


@dataclass
class QualityIndicator:
    """Výsledok analýzy kvality vstupných dokumentov."""
    files_count:        int
    total_size_kb:      int
    total_text_chars:   int        # celková dĺžka extrahovaného textu (vzorky)
    keyword_hits:       int        # počet detekovaných FMEA-relevantných výrazov
    detected_types:     dict       # napr. {"PDF": 4, "DOCX": 1}
    detected_sections:  dict       # napr. {"workflow": True, "control_plan": False}
    section_hit_counts: dict       # počet zhôd na sekciu (bohatosť kontextu)
    score:              int        # 0–100
    score_label:        str        # "Vynikajúce" / "Veľmi dobré" / "Dobré" / "Postačujúce" / "Slabé"
    score_color:        str        # hex farba pre vizualizáciu
    recommendations:    list[str]  # zoznam odporúčaní


# ── Cenový odhad ───────────────────────────────────────────────────────────────

def _read_file_chars(path: Path) -> int:
    """
    Vráti odhadovaný počet znakov textového obsahu súboru.

    Pre PDF/DOCX/XLSX použijeme heuristiku podľa veľkosti súboru:
    - PDF:  ~30 % súboru je text (zvyšok je metadata, fonty, obrázky)
    - DOCX: ~25 % súboru je text (XML overhead + štýly)
    - XLSX: ~30 % súboru je text
    - TXT:  ~95 % súboru je text (priamy obsah)
    """
    try:
        size_bytes = path.stat().st_size
    except (OSError, AttributeError):
        return 0

    suffix = path.suffix.lower()
    text_ratio = {
        ".txt":  0.95,
        ".md":   0.95,
        ".pdf":  0.30,
        ".docx": 0.25,
        ".xlsx": 0.30,
    }.get(suffix, 0.20)

    return int(size_bytes * text_ratio)


def estimate_cost(file_paths: list[Path], model: str = "claude-opus-4-7") -> CostEstimate:
    """
    Odhadne cenu za jeden kompletný beh PFMEA Tool.

    Výpočet:
      1. Suma znakov vstupných dokumentov → tokeny
      2. Odhad krokov procesu na základe veľkosti vstupu
      3. Pre každý krok: kontext (~5000 znakov) + system prompt + output
      4. Plus 2 úvodné AI volania (názov procesu + extrakcia krokov)
    """
    pricing = MODEL_PRICING.get(model, DEFAULT_PRICING)

    # 1) Vstupné znaky a tokeny
    total_chars = sum(_read_file_chars(Path(p)) for p in file_paths)
    base_input_tokens = int(total_chars / CHARS_PER_TOKEN_SLOVAK)

    # 2) Odhad počtu krokov (min 5, max 30)
    estimated_steps = max(5, min(30, int(total_chars / 1000 * STEPS_PER_1K_CHARS)))

    # 3) Vstupné tokeny pre AI volania:
    #    a) infer_process_name_from_ai – posiela prvých 6000 znakov
    #    b) extract_process_steps – posiela max 16000 znakov + system prompt
    #    c) generate_fmea_for_step – pre každý krok kontext ~5000 znakov + system

    name_inference_tokens = int(min(6000, total_chars) / CHARS_PER_TOKEN_SLOVAK) + SYSTEM_PROMPT_TOKENS
    steps_extraction_tokens = int(min(16000, total_chars) / CHARS_PER_TOKEN_SLOVAK) + SYSTEM_PROMPT_TOKENS
    per_step_input_tokens = int(5000 / CHARS_PER_TOKEN_SLOVAK) + SYSTEM_PROMPT_TOKENS

    total_input_tokens = (
        name_inference_tokens
        + steps_extraction_tokens
        + estimated_steps * per_step_input_tokens
    )

    # 4) Output tokeny:
    #    a) názov procesu: ~30 tokenov
    #    b) extrakcia krokov: ~50 tokenov na krok
    #    c) FMEA položky: 2.5 položiek na krok × 500 tokenov
    estimated_items = int(estimated_steps * ITEMS_PER_STEP_AVG)
    output_tokens = (
        30
        + estimated_steps * 50
        + estimated_items * TOKENS_PER_FMEA_ITEM
    )

    # 5) Cena
    cost_input_usd  = (total_input_tokens / 1_000_000) * pricing["input"]
    cost_output_usd = (output_tokens      / 1_000_000) * pricing["output"]
    cost_total_usd  = cost_input_usd + cost_output_usd

    # USD → EUR (orientačný kurz; pre presnejší použiť API)
    EUR_PER_USD = 0.92
    cost_total_eur = cost_total_usd * EUR_PER_USD

    return CostEstimate(
        total_input_chars  = total_chars,
        total_input_tokens = total_input_tokens,
        estimated_steps    = estimated_steps,
        estimated_items    = estimated_items,
        output_tokens      = output_tokens,
        cost_input_usd     = round(cost_input_usd,  4),
        cost_output_usd    = round(cost_output_usd, 4),
        cost_total_usd     = round(cost_total_usd,  3),
        cost_total_eur     = round(cost_total_eur,  3),
        model              = model,
        files_count        = len(file_paths),
    )


# ── Analýza kvality vstupov ────────────────────────────────────────────────────

# Kľúčové slová ktoré indikujú prítomnosť konkrétnych typov FMEA podkladov
SECTION_KEYWORDS = {
    "pracovny_postup": [
        # Pracovné postupy a inštrukcie
        "pracovný postup", "pracovny postup", "work instruction", "wi ",
        "procesný tok", "process flow", "postup práce",
        # Návody na montáž, inštaláciu, prevádzku, údržbu (rovnako relevantné)
        "návod na montáž", "navod na montaz", "návod k montáži",
        "návod na inštaláciu", "navod na instalaciu", "installation manual",
        "návod na obsluhu", "navod na obsluhu", "návod na použitie",
        "návod na prevádzku", "návod na údržbu", "operating manual",
        "uvedenie do prevádzky", "uvedenie do prevadzky", "commissioning",
        "montážny postup", "montaz", "montáž",
        # Procesné kroky a operácie
        "operácia", "operacia", "operation",
        "krok ", "úkon", "ukon", "step ",
        "fáza", "faza", "phase",
        "procedúra", "procedura", "procedure",
        # Demontáž / likvidácia (typický záver návodu)
        "demontáž", "demontaz", "disassembly",
    ],
    "kontrolny_plan": [
        # Kontrolné plány
        "kontrolný plán", "kontrolny plan", "control plan",
        # Kontrolné činnosti a parametre
        "kontrola", "skontrolujte", "kontrolovať",
        "meranie", "merací", "tolerancia", "spc",
        "frekvencia kontroly", "kritická charakteristika",
        # Špecifikácie a parametre
        "krútiaci moment", "krutiaci moment", "torque", "nm)",
        "menovitý", "menovity", "menovité",
        "technické údaje", "technicke udaje", "technical data",
        "parametre", "parameter", "parametrov",
        "rozmer", "rozmery", "dimension",
        # Bezpečnostné kontroly
        "bezpečnostné", "bezpecnostne", "safety check",
        "ochranné zariadenia", "ochranne zariadenia",
        "preskúšanie", "preskusanie", "test report",
    ],
    "zaznamy_nezhod": [
        # Záznamy o nezhodách
        "nezhoda", "nezhody", "nonconformance", "non-conformance",
        "ncr", "8d report", "ppm",
        "reklamácia", "reklamacia", "claim", "complaint",
        # Chyby a poruchy
        "chyba", "chyby", "porucha", "poruchy", "fault", "failure",
        "vada", "vady", "defect",
        "poškodenie", "poskodenie", "damage", "poškodený",
        # Zlyhania a riziká
        "zlyhanie", "zlyhanie", "porušenie", "porušenia",
        "neúmyselný", "neumyselny", "neočakávaný", "neocakavany",
        "nebezpečenstvo", "nebezpecenstvo", "hazard", "risk",
        # Výstrahy ktoré naznačujú možné chyby
        "varovanie", "warning", "upozornenie",
        "pozor", "caution", "danger",
    ],
    "udrzba_kalibrácia": [
        # Údržba a kalibrácia
        "údržba", "udrzba", "maintenance", "servis",
        "kalibrácia", "kalibracia", "calibration",
        "preventívna", "preventivna", "preventive",
        "plán údržby", "plan udrzby", "maintenance plan",
        # Mazanie, výmena, opravy
        "mazanie", "namažte", "namazat", "lubrication",
        "výmena", "vymena", "replacement",
        "oprava", "opravy", "repair",
        # Životnosť a opotrebenie
        "životnosť", "zivotnost", "service life",
        "opotrebenie", "opotrebovanie", "wear",
        "cyklus", "cyklov", "cycles",
    ],
    "fmea_existujuca": [
        "fmea", "rpn", "severity", "occurrence", "detection",
        "význam", "vyznam", "výskyt", "vyskyt", "odhalenie",
        "možná chyba", "mozna chyba", "failure mode",
        "klasifikácia", "klasifikacia",
    ],
}


def _read_file_text_sample(path: Path, max_chars: int = 5000) -> str:
    """
    Načíta vzorku textu zo súboru pre detekciu sekcií.

    Optimalizácie:
    - PDF: parsuje max prvých 5 strán (väčšie PDF sú často projektové dokumenty
      s desiatkami strán, ale kľúčová terminológia je vždy na prvých stranách)
    - DOCX: zastaví sa pri max_chars, neiteruje cez celý dokument
    - XLSX: read_only mode + zastavenie po prvom hárku ak je už nazbieraný text
    """
    suffix = path.suffix.lower()

    try:
        if suffix in (".txt", ".md"):
            return path.read_text(encoding="utf-8", errors="ignore")[:max_chars]

        if suffix == ".pdf":
            from pypdf import PdfReader
            reader = PdfReader(str(path))
            text_parts = []
            char_count = 0
            # Limitujeme na prvých 5 strán – aj veľké 50MB PDF takto trvá <1s
            for page_idx, page in enumerate(reader.pages):
                if page_idx >= 5:
                    break
                page_text = page.extract_text() or ""
                text_parts.append(page_text)
                char_count += len(page_text)
                if char_count >= max_chars:
                    break
            return "\n".join(text_parts)[:max_chars]

        if suffix == ".docx":
            from docx import Document
            doc = Document(str(path))
            text_parts = []
            char_count = 0
            for para in doc.paragraphs:
                if para.text.strip():
                    text_parts.append(para.text)
                    char_count += len(para.text)
                    if char_count >= max_chars:
                        break
            return "\n".join(text_parts)[:max_chars]

        if suffix == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(filename=str(path), data_only=True, read_only=True)
            text_parts = []
            char_count = 0
            try:
                for sheet_name in wb.sheetnames:
                    text_parts.append(sheet_name)
                    for row in wb[sheet_name].iter_rows(values_only=True):
                        cells = [str(v).strip() for v in row if v is not None]
                        if cells:
                            line = " ".join(cells)
                            text_parts.append(line)
                            char_count += len(line)
                            if char_count >= max_chars:
                                return "\n".join(text_parts)[:max_chars]
            finally:
                wb.close()
            return "\n".join(text_parts)[:max_chars]

    except Exception:
        return ""

    return ""


def analyze_quality(file_paths: list[Path]) -> QualityIndicator:
    """
    Analyzuje vstupné súbory a vráti kvalitatívny ukazovateľ s odporúčaniami.

    Skóre 0–100 sa počíta na základe:
    - Počtu súborov  (≥3 = optimum)
    - Diverzity formátov  (PDF + DOCX + XLSX = lepšie)
    - Detekovaných typov sekcií  (každá detekovaná sekcia +20 bodov)
    """
    if not file_paths:
        return QualityIndicator(
            files_count=0, total_size_kb=0,
            total_text_chars=0, keyword_hits=0,
            detected_types={}, detected_sections={},
            section_hit_counts={},
            score=0, score_label="Žiadne dokumenty",
            score_color="#dc2626",
            recommendations=["Pridajte aspoň 2 vstupné dokumenty pre kvalitnú analýzu."],
        )

    paths = [Path(p) for p in file_paths]

    # Typy súborov
    detected_types: dict = {}
    total_size = 0
    for p in paths:
        ext = p.suffix.upper().lstrip(".")
        detected_types[ext] = detected_types.get(ext, 0) + 1
        try:
            total_size += p.stat().st_size
        except OSError:
            pass

    # Detekcia sekcií + meranie dĺžky textu a bohatosti kontextu
    # Načítame väčšiu vzorku textu (8000 znakov) pre presnejšiu analýzu obsahu
    combined_text = ""
    total_text_chars = 0
    for p in paths[:8]:  # max 8 súborov pre rýchlosť
        sample = _read_file_text_sample(p, max_chars=8000)
        combined_text += sample + "\n"
        total_text_chars += len(sample)
    combined_text_lower = combined_text.lower()

    # Pre každú sekciu spočítame koľko kľúčových slov sa našlo (bohatosť kontextu)
    detected_sections = {}
    section_hit_counts = {}
    for section_key, keywords in SECTION_KEYWORDS.items():
        hits = sum(1 for kw in keywords if kw in combined_text_lower)
        detected_sections[section_key] = hits > 0
        section_hit_counts[section_key] = hits

    # Celkový počet detekovaných FMEA-relevantných výrazov (z všetkých sekcií)
    total_keyword_hits = sum(section_hit_counts.values())

    # ── Výpočet skóre ──────────────────────────────────────────────────────────
    # Skóre 0–100 sa skladá z piatich faktorov:
    #   1. Počet súborov          (max 25)
    #   2. Diverzita formátov     (max 10)
    #   3. Pokrytie typovými sekciami  (max 35)
    #   4. Dĺžka reálneho obsahu  (max 15)
    #   5. Bohatosť FMEA kontextu (max 15)
    # + bonusy: existujúca FMEA (5), veľký objem (5)
    score = 0
    n = len(paths)

    # 1) Počet súborov – odmeňujeme 3+ dokumentov (max 25)
    if n >= 5:
        score += 25
    elif n == 4:
        score += 22
    elif n == 3:
        score += 19
    elif n == 2:
        score += 12
    else:
        score += 5

    # 2) Diverzita formátov (max 10)
    n_types = len(detected_types)
    if n_types >= 3:
        score += 10
    elif n_types == 2:
        score += 7
    else:
        score += 4

    # 3) Pokrytie typovými sekciami – ÁNO/NIE základ (max 35)
    if detected_sections.get("pracovny_postup"):
        score += 12
    if detected_sections.get("kontrolny_plan"):
        score += 10
    if detected_sections.get("zaznamy_nezhod"):
        score += 9
    if detected_sections.get("udrzba_kalibrácia"):
        score += 4

    # 4) Dĺžka reálneho obsahu – ako veľa textu AI dostane k dispozícii (max 15)
    # Hodnota total_text_chars je suma znakov z prvých 8000 chars každého súboru.
    # Plný kredit dostane analýza s 30k+ znakov reálneho textu (=zhruba 8 strán A4).
    if total_text_chars >= 30000:
        score += 15
    elif total_text_chars >= 18000:
        score += 12
    elif total_text_chars >= 10000:
        score += 9
    elif total_text_chars >= 5000:
        score += 6
    elif total_text_chars >= 2000:
        score += 3
    else:
        score += 1

    # 5) Bohatosť FMEA kontextu – koľko relevantných pojmov sa v texte vyskytuje (max 15)
    # Vychádza z celkového počtu zhôd kľúčových slov naprieč všetkými sekciami.
    # Bohatý text obsahuje desiatky technických pojmov (operácia, kontrola, tolerancia...)
    if total_keyword_hits >= 25:
        score += 15
    elif total_keyword_hits >= 18:
        score += 12
    elif total_keyword_hits >= 12:
        score += 9
    elif total_keyword_hits >= 7:
        score += 6
    elif total_keyword_hits >= 3:
        score += 3
    else:
        score += 0

    # 6) Bonus za prítomnosť existujúcej FMEA (referenčný materiál) – max 5
    if detected_sections.get("fmea_existujuca"):
        score += 5

    # 7) Bonus za veľký objem dát (komplexný proces, dostatok kontextu) – max 5
    total_size_mb = sum(
        p.stat().st_size for p in paths if p.exists()
    ) / (1024 * 1024)
    if total_size_mb >= 10:
        score += 5
    elif total_size_mb >= 3:
        score += 3

    # Cap na 100
    score = min(100, score)

    # ── Klasifikácia skóre ─────────────────────────────────────────────────────
    if score >= 85:
        label, color = "Vynikajúce", "#16a34a"
    elif score >= 70:
        label, color = "Veľmi dobré", "#22c55e"
    elif score >= 55:
        label, color = "Dobré", "#65a30d"
    elif score >= 40:
        label, color = "Postačujúce", "#f59e0b"
    elif score >= 25:
        label, color = "Slabé", "#ef4444"
    else:
        label, color = "Nedostatočné", "#dc2626"

    # ── Odporúčania – kontextové podľa skóre a slabých faktorov ──────────────
    recommendations = []

    # Kritické chýbajúce sekcie – iba ak skóre nie je vynikajúce
    if score < 85:
        if not detected_sections.get("pracovny_postup"):
            recommendations.append(
                "Pridajte pracovný postup alebo procesný tok – pomôže AI lepšie identifikovať kroky procesu."
            )
        if not detected_sections.get("kontrolny_plan"):
            recommendations.append(
                "Pridajte kontrolný plán – umožní presnejšie hodnotenie metód odhalenia (D)."
            )
        if not detected_sections.get("zaznamy_nezhod"):
            recommendations.append(
                "Pridajte záznamy o nezhodách – zlepší hodnotenie pravdepodobnosti výskytu (O)."
            )

    # Malý počet dokumentov
    if n < 3:
        recommendations.append(
            "Odporúčame použiť aspoň 3 dokumenty rôznych typov pre komplexnejší pohľad na proces."
        )

    # Krátke dokumenty – AI nemá dosť kontextu
    if total_text_chars < 5000 and n >= 2:
        recommendations.append(
            "Vstupné dokumenty obsahujú málo textu – AI bude pracovať s obmedzeným kontextom. "
            "Pridajte podrobnejšie pracovné postupy alebo technickú dokumentáciu."
        )

    # Slabá bohatosť kontextu – obsah nie je dostatočne FMEA-relevantný
    if total_keyword_hits < 7 and n >= 2:
        recommendations.append(
            "Obsah dokumentov má nízky podiel FMEA-relevantnej terminológie. "
            "Skontrolujte, či dokumenty popisujú výrobný proces a jeho kontroly."
        )

    # Pozitívna spätná väzba pri absencii výhrad
    if not recommendations:
        if score >= 85:
            recommendations.append(
                "Vstupné dokumenty pokrývajú všetky kľúčové oblasti – analýza bude veľmi kvalitná."
            )
        else:
            recommendations.append(
                "Vstupné dokumenty pokrývajú kľúčové oblasti – analýza by mala byť kvalitná."
            )

    # Pri vynikajúcom skóre doplň konkrétne metriky obsahu
    if score >= 85:
        sections_count = sum(1 for v in detected_sections.values() if v)
        # Koľko strán A4 textu má AI k dispozícii (priemerne 3500 znakov / strana)
        approx_pages = max(1, total_text_chars // 3500)
        recommendations.append(
            f"Detekovaných {sections_count} z 5 typových sekcií, "
            f"≈ {approx_pages} strán textu, {total_keyword_hits} odborných pojmov – "
            f"AI má bohatý kontext pre analýzu."
        )

    return QualityIndicator(
        files_count        = len(paths),
        total_size_kb      = total_size // 1024,
        total_text_chars   = total_text_chars,
        keyword_hits       = total_keyword_hits,
        detected_types     = detected_types,
        detected_sections  = detected_sections,
        section_hit_counts = section_hit_counts,
        score              = score,
        score_label        = label,
        score_color        = color,
        recommendations    = recommendations,
    ) 